#!/usr/bin/env python
# -*- coding: utf-8 -*-

import logging
import datetime
import time
import tempfile
import openpyxl
from pathlib import Path
from re import search as re_search
from Base import TPL_FORMAT

logger = logging.getLogger('xxx')

def string_none(x=None):
    if x is None:
        return ''
    else:
        return str(x)

def IsOraInstalled():
    try:
        import cx_Oracle
        global cx_Oracle
        return True
    except Exception as zz:
        logger.error("{0} - ERROR: {1}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), zz))
        return False

def ConnDb(db_type='ora', alias_db=None):
    if db_type=='ora':
        try:
            new_dsn = cx_Oracle.makedsn(host=alias_db['host'], port=alias_db['port'], service_name=alias_db['sid'])
            connection = cx_Oracle.connect(alias_db['user'], alias_db['passw'], new_dsn)
            connection.autocommit = False
            logger.info("> connected to Oracle, client {0}".format('.'.join([str(i) for i in cx_Oracle.clientversion()])))
            return connection
        except cx_Oracle.DatabaseError as exc:
            error, = exc.args
            logger.error('{0}\n'.format(error.message.strip()))
            return False
    else:
        logger.error("Only ORACLE connection available in this version")
        return False

def CheckSQLSafety(txt=''):
    txt_temp = txt.lower().strip()
    if ('alter' in txt_temp or 'insert' in txt_temp or 'delete' in txt_temp or 'drop' in txt_temp or 'update' in txt_temp):
        logger.error('\nACHTUNG !!! - Dangerous SQL !!!\n')
        return False
    else:
        return txt

def xxxxxxxxxx(cfg=None, arg=None, cursor=None, sql_prepared=None, param=None, result_path=None):
    try:
        logger.info('  > start - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        if param:
            tt = (', ', ')')
        else:
            param = {}
            tt = (')', '')
        xx = ''.join(['cursor.execute(sql_prepared{0}'.format(tt[0]), ''.join([''.join([param[i][0], "='", param[i][1], "', "]) for i in param.keys()])[:-2], tt[1]])
        for_eval = {'cursor': cursor, 'sql_prepared': sql_prepared}
        eval(xx, for_eval)
        file_out = openpyxl.Workbook(optimized_write = True, guess_types=False)
        scount = 0
        file_out_sheet = file_out.create_sheet(title=''.join(['sheet', str(scount)]))
        file_out_sheet.append([str(i[0]) for i in cursor.description])
        count = 0
        count2= 0
        for data_row in cursor:
            if (count % 10000) == 0 and count != 0:
                logger.info('  ...{0}'.format(str(count)))
            if count < int(cfg[arg.section]['bound']):
                tmp_list = [str(k).replace('None', '') for k in data_row]
                file_out_sheet.append(tmp_list)
            else:
                scount = scount + 1
                file_out_sheet = file_out.create_sheet(title=''.join(['sheet', str(scount)]))
                file_out_sheet.append([str(i[0]) for i in cursor.description])
                tmp_list = [str(k).replace('None', '') for k in data_row]
                file_out_sheet.append(tmp_list)
                count = 0
            count = count + 1
            count2= count2+ 1
        if arg.write_count is True:
            file_out_name = Path(''.join([str(result_path),'_','_'.join([param[i][1] for i in param]).replace(':','_') , '_U', str(count2), '.xlsx']))
        else:
            file_out_name = Path(''.join([str(result_path),'_','_'.join([param[i][1] for i in param]).replace(':','_') , '_U.xlsx']))
        file_out.save(str(file_out_name))
        if cursor.rowcount != 0:
            logger.info('  > to save: {0}'.format(count2))
        else:
            logger.warning('  > to save: {0}'.format(count2))
            if arg.save_empty is False:
                file_out_name.unlink()
        logger.info('  > stop  - {0}\n'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        return str(file_out_name)
    except cx_Oracle.DatabaseError as exc1:
        error, = exc1.args
        logger.error('{0}\n'.format(error.message.strip()))
        return str('')

def wSimple(arg=None, cfg=None, sql_root=None, res_root=None):
    logger.info("Unload - simple")
    sql_list = list(sql_root.glob('**/*.sql'))
    if not res_root.exists():
        res_root.mkdir(parents=True)
    if sql_list:
        dd = None
        if arg.db is None:
            dd = cfg[cfg[str(arg.section)]['db_alias']]
            logger.info(str(cfg[str(arg.section)]['db_alias']))
        else:
            dd = cfg[arg.db]
            logger.info(arg.db)
        cnn = ConnDb('ora', dd)
        if cnn:
            cursor = cnn.cursor()
            sr = Path(str(cfg[arg.section]['saved_filenames']))
            if not sr.is_dir() and not sr.parent.exists():
                sr.parent.mkdir(parents=True)
            if arg.save_filenames is True and not sr.is_dir():
                ftemp = sr.open('w')
            else:
                ftemp = tempfile.TemporaryFile('w')
            for i in sql_list:
                logger.info('-> {0}'.format(Path(i).name))
                sql_content = CheckSQLSafety(i.open('r').read())
                file_out_prefix = None
                if arg.use_src_subfolders:
                    file_out_prefix = res_root.joinpath(*i.parts[1:]).parent
                else:
                    file_out_prefix = res_root
                if not file_out_prefix.exists():
                    file_out_prefix.mkdir(parents=True)
                if sql_content:
                    zx = xxxxxxxxxx(cfg, arg, cursor, sql_content, None, file_out_prefix/Path(i.name).stem)
                    ftemp.write(zx+'\n')
                else:
                    continue
            ftemp.close()
            cnn.close()
            return True
    else:
        logger.error('ERROR: No SQL in {0} or unable to create {1}\n'.format(sql_root, res_root))
        return False

def wParametric(arg=None, cfg=None, sql_root=None, res_root=None):
    logger.info("Unload - parametric")
    sql_list = list(sql_root.glob('**/*.sql'))
    if not res_root.exists():
        res_root.mkdir(parents=True)
    if sql_list:
        dd = None
        if arg.db is None:
            dd = cfg[cfg[str(arg.section)]['db_alias']]
            logger.info(str(cfg[str(arg.section)]['db_alias']))
        else:
            dd = cfg[arg.db]
            logger.info(arg.db)
        cnn = ConnDb('ora', dd)
        if cnn:
            cursor = cnn.cursor()
            sr = Path(str(cfg[arg.section]['saved_filenames']))
            if not sr.is_dir() and not sr.parent.exists():
                sr.parent.mkdir(parents=True)
            if arg.save_filenames is True and not sr.is_dir():
                ftemp = sr.open('w')
            else:
                ftemp = tempfile.TemporaryFile('w')
            for i in sql_list:
                sql_content = CheckSQLSafety(i.open('r').read())
                file_out_prefix = None
                if arg.use_src_subfolders:
                    file_out_prefix = res_root.joinpath(*i.parts[1:]).parent
                else:
                    file_out_prefix = res_root
                if not file_out_prefix.exists():
                    file_out_prefix.mkdir(parents=True)
                for j in TPL_FORMAT[cfg[arg.section]['param']]:
                    logger.info('-> {0} - {1}'.format(Path(i.name), j))
                    if sql_content:
                        zx = xxxxxxxxxx(cfg, arg, cursor, sql_content, j, file_out_prefix/Path(i.name).stem)
                        ftemp.write(zx+'\n')
                    else:
                        continue
            ftemp.close()
            cnn.close()
            return True
    else:
        logger.error('ERROR: No SQL in {0} or unable to create {1}\n'.format(sql_root, res_root))
        return False

def wAllPvd(arg=None, cfg=None, sql_root=None, res_root=None):
    logger.info("Unload - PVD")
    sql_list = list(sql_root.glob('**/*.sql'))
    if not res_root.exists():
        res_root.mkdir(parents=True)
    if sql_list:
        sr = Path(str(cfg[arg.section]['saved_filenames']))
        if not sr.is_dir() and not sr.parent.exists():
            sr.parent.mkdir(parents=True)
        if arg.save_filenames is True and not sr.is_dir():
            ftemp = sr.open('w')
        else:
            ftemp = tempfile.TemporaryFile('w')
        for j in sorted(cfg.keys()):
            if cfg[j]['type']=='pvd':
                logger.info(j)
                cnn = ConnDb('ora', cfg[j])
                if cnn:
                    cursor = cnn.cursor()
                    for i in sql_list:
                        logger.info('-> {0}'.format(Path(i.name)))
                        sql_content = CheckSQLSafety(i.open('r').read())
                        file_out_prefix = None
                        if arg.use_src_subfolders:
                            file_out_prefix = res_root.joinpath(*i.parts[1:]).parent
                        else:
                            file_out_prefix = res_root
                        if not file_out_prefix.exists():
                            file_out_prefix.mkdir(parents=True)
                        if sql_content:
                            zx = xxxxxxxxxx(cfg, arg, cursor, sql_content, None, file_out_prefix/''.join([Path(i.name).stem, '_', j]))
                            ftemp.write(zx+'\n')
                        else:
                            continue
                    cnn.close()
        ftemp.close()
        return True
    else:
        logger.error('ERROR: No SQL in {0} or unable to create {1}\n'.format(sql_root, res_root))
        return False
        
def wUnloadByXlsxCol(arg=None, cfg=None, sql_root=None, res_root=None):
    logger.info("Unload_By_Xlsx_Col (using UNLOAD_ACTIONS_LIST)")
    try:
        xls_name_list = list(sql_root.glob('**/*.xlsx'))
        if (len(xls_name_list) == 0):
            logger.error("{0}: {1} {2}".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), "No XLSX in", sql_root))
            return False
        else:
            if not res_root.exists():
                res_root.mkdir(parents=True)
            selected_action = cfg['UNLOAD_ACTIONS_LIST'][str(arg.section)]
            sql_content = CheckSQLSafety(open(selected_action, 'r').read())
            if sql_content:
                dd = None
                if arg.db is None:
                    dd = cfg[cfg['UNLOAD_ACTIONS_LIST']['db_alias']]
                    logger.info('{0} / {1} / DB: {2}'.format(str(arg.section), 'SQL - OK', cfg['UNLOAD_ACTIONS_LIST']['db_alias']))
                else:
                    dd = cfg[arg.db]
                    logger.info('{0} / {1} / DB: {2}'.format(str(arg.section), 'SQL - OK', str(arg.db)))
                cnn = ConnDb('ora', dd)
                if cnn:
                    curs = cnn.cursor()
                    sr = Path(str(cfg['UNLOAD_ACTIONS_LIST']['saved_filenames']))
                    if not sr.is_dir() and not sr.parent.exists():
                        sr.parent.mkdir(parents=True)
                    if arg.save_filenames is True and not sr.is_dir():
                        ftemp = sr.open('w')
                    else:
                        ftemp = tempfile.TemporaryFile('w')
                    try:
                        for i in xls_name_list:
                            try:
                                wb = openpyxl.Workbook(optimized_write = True, guess_types=False)
                                ws = wb.create_sheet()
                                file_in = str(i)
                                file_out_prefix = None
                                if arg.use_src_subfolders:
                                    file_out_prefix = res_root.joinpath(*i.parts[1:]).parent
                                else:
                                    file_out_prefix = res_root
                                if not file_out_prefix.exists():
                                    file_out_prefix.mkdir(parents=True)
                                file_out = file_out_prefix/''.join([Path(i.name).stem, '_END.xlsx'])
                                load_book = openpyxl.load_workbook(file_in, read_only=True, use_iterators=True)
                                load_sheet = load_book.active
                                agg = []
                                r_count = 0
                                logger.info(' -> file : {0}'.format(i.name))
                                logger.info('  > start: {0}'.format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S')))
                                try:
                                    for r in load_sheet.rows:
                                        rs = []
                                        agg = [string_none(cc.value) for cc in r]
                                        patt = '^.*$'
                                        if agg[int(cfg['UNLOAD_ACTIONS_LIST']['kolonka'])] is not None:
                                            z = str(agg[int(cfg['UNLOAD_ACTIONS_LIST']['kolonka'])]).strip()
                                            if not re_search(patt, z):
                                                ws.append(agg)
                                                r_count += 1
                                            else:
                                                curs.execute(sql_content, S1 = z)
                                                is_not_empty = 0
                                                abc = curs.fetchall()
                                                if len(abc) != 0:
                                                    is_not_empty = 1
                                                    for ii in abc:
                                                        a = list(ii)
                                                        rs = agg + [string_none(x) for x in a]
                                                        ws.append(rs)
                                                if is_not_empty == 0:
                                                    ws.append(agg)
                                                r_count += 1
                                                time.sleep(0.2)
                                            logger.info('    {0}{1}'.format(str(r_count).ljust(7), z))
                                        else:
                                            ws.append(agg)
                                            r_count += 1
                                    wb.save(str(file_out))
                                    ftemp.write(str(file_out)+'\n')
                                    logger.info('  > stop : {0}\n'.format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S')))
                                except Exception as eee:
                                    logger.error("{0}: {1}\n{2}\n{3}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), i, agg[int(cfg['UNLOAD_ACTIONS_LIST']['kolonka'])], eee))
                                    continue
                            except Exception as eee:
                                logger.error("{0}: {1}\n{2}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), i, eee))
                                continue
                    finally:
                        logger.info('> close connection')
                        cnn.close()
                        ftemp.close()
                else:
                    return False
            else:
                return False
    except Exception as zz:
        logger.error("{0} - ERROR: {1}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), zz))
        return False

def Unloading(arg=None, cfg=None):
    if arg and IsOraInstalled() and cfg:
        if arg.action == 'unload' and arg.type == 'by_xlsx_col':
            sql_root = Path(str(cfg['UNLOAD_ACTIONS_LIST']['source']))
            res_root = Path(str(cfg['UNLOAD_ACTIONS_LIST']['destination']))
        else:
            sql_root = Path(str(cfg[str(arg.section)]['source']))
            res_root = Path(str(cfg[str(arg.section)]['destination']))
        subtypes = {\
            'simple': wSimple,
            'parametric': wParametric,
            'allpvd': wAllPvd,
            'by_xlsx_col': wUnloadByXlsxCol
            }
        subtypes[arg.type](arg, cfg, sql_root, res_root)
        logger.info('END\n-----------\n\n')
        return True
    else:
        return False

if __name__ == '__main__':
    pass
