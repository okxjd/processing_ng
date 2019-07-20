#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
import logging
import datetime
import openpyxl
import tempfile
from pathlib import Path

logger = logging.getLogger('xxx')

def string_none(x=None):
    if x is None:
        return ''
    else:
        return str(x)

def XLSX_Join(arg=None, cfg=None):
    logger.info('XLSX_Join')
    file_dir   = Path(str(cfg[str(arg.section)]['source']))
    result_dir = Path(str(cfg[str(arg.section)]['destination']))
    xlsx_list  = list(file_dir.glob('**/*.xlsx'))
    if not result_dir.exists():
        result_dir.mkdir(parents=True)
    if xlsx_list:
        logger.info(' > start - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        sr = Path(str(cfg[arg.section]['saved_filenames']))
        if arg.save_filenames is True and not sr.is_dir():
            ftemp = sr.open('w')
        else:
            ftemp = tempfile.TemporaryFile('w')
        wb = openpyxl.Workbook(optimized_write=True, guess_types=False, data_only=True)
        scount = 0
        ws = wb.create_sheet(title=''.join(['sheet', str(scount)]))
        cnt = 0
        for i in xlsx_list:
            try:
                logger.info('  > {0}'.format(Path(i).name))
                r_book = openpyxl.load_workbook(str(i), read_only=True, use_iterators=True)
                r_sheet = r_book.active
                tmp_row = []
                for rrr in r_sheet.rows:
                    if (cnt % 10000) == 0 and cnt != 0:
                        logger.info('  ...{0}'.format(str(cnt)))
                    tmp_row = [string_none(i.value) for i in rrr]
                    cnt = cnt + 1
                    ws.append(tmp_row)
            except Exception as ee:
                logger.error("{0} - ERROR: {1}\n{2}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), i, ee))
        logger.info(' > to save: {0}'.format(str(cnt)))
        if arg.write_count:
            result_file = str(Path(result_dir).joinpath(''.join([Path(Path(xlsx_list[0]).name).stem, '_J', str(cnt), '.xlsx'])))
        else:
            result_file = str(Path(result_dir).joinpath(''.join([Path(Path(xlsx_list[0]).name).stem, '_J.xlsx'])))
        wb.save(result_file)
        ftemp.write(result_file+'\n')
        ftemp.close()
        logger.info(' > stop  - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        return True
    else:
        logger.error('ERROR: No XLSX in {0} or unable to create {1}'.format(str(file_dir), str(result_dir)))
        return False

def XLSX_Split_by_rowcount(arg=None, cfg=None):
    logger.info('XLSX_Split_by_rowcount')
    file_dir   = Path(str(cfg[str(arg.section)]['source']))
    result_dir = Path(str(cfg[str(arg.section)]['destination']))
    xlsx_list  = list(file_dir.glob('**/*.xlsx'))
    split_num  = int(cfg[str(arg.section)]['split_row_cnt'])
    logger.info(' > split by {0} rows'.format(split_num))
    if not result_dir.exists():
        result_dir.mkdir(parents=True)
    if xlsx_list:
        logger.info(' > start - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        sr = Path(str(cfg[arg.section]['saved_filenames']))
        if arg.save_filenames is True and not sr.is_dir():
            ftemp = sr.open('w')
        else:
            ftemp = tempfile.TemporaryFile('w')
        for i in xlsx_list:
            try:
                logger.info('  > {0}'.format(Path(i).name))
                r_book = openpyxl.load_workbook(str(i), read_only=True, use_iterators=True)
                r_sheet = r_book.active
                tmp_row = []
                cnt = 0
                fc = 0
                scount = 0
                header = []
                wb = openpyxl.Workbook(optimized_write=True, guess_types=False, data_only=True)
                ws = wb.create_sheet(title=''.join(['sheet', str(scount)]))
                for rrr in r_sheet.rows:
                    if cnt == 0:
                        header = [string_none(i.value) for i in rrr]
                        ws.append(header)
                        cnt = cnt + 1
                    else:
                        tmp_row = [string_none(i.value) for i in rrr]
                        if cnt < split_num:
                            ws.append(tmp_row)
                            cnt = cnt + 1
                        else:
                            result_file = str(Path(result_dir).joinpath(''.join([Path(Path(i).name).stem, '_', str(fc).zfill(3), '.xlsx'])))
                            wb.save(result_file)
                            logger.info('   -> {0}'.format(''.join([Path(Path(i).name).stem, '_', str(fc).zfill(3), '.xlsx'])))
                            wb = openpyxl.Workbook(optimized_write=True, guess_types=False, data_only=True)
                            ws = wb.create_sheet(title=''.join(['sheet', str(scount)]))
                            ws.append(header)
                            ws.append(tmp_row)
                            fc = fc + 1
                            cnt = 1
                result_file = str(Path(result_dir).joinpath(''.join([Path(Path(i).name).stem, '_', str(fc).zfill(3), '.xlsx'])))
                wb.save(result_file)
                ftemp.write(result_file+'\n')
                logger.info('   -> {0}\n'.format(''.join([Path(Path(i).name).stem, '_', str(fc).zfill(3), '.xlsx'])))
            except Exception as ee:
                logger.error("{0} - ERROR: {1}\n{2}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), i, ee))
                continue
        ftemp.close()
        logger.info(' > stop  - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        return True
    else:
        logger.error('ERROR: No XLSX in {0} or unable to create {1}'.format(str(file_dir), str(result_dir)))
        return False
        
def XLSX_Split_by_mask(arg=None, cfg=None):
    logger.info('XLSX_Split_by_mask')
    file_dir   = Path(str(cfg[str(arg.section)]['source']))
    result_dir = Path(str(cfg[str(arg.section)]['destination']))
    xlsx_list  = list(file_dir.glob('**/*.xlsx'))
    col_match  = int(cfg[str(arg.section)]['split_col_for_search'])
    split_mask = str(cfg[str(arg.section)]['split_mask'])
    logger.info(' > column: {0}, mask: {1}'.format(col_match, split_mask))
    mask = re.compile(split_mask)
    if not result_dir.exists():
        result_dir.mkdir(parents=True)
    if xlsx_list:
        logger.info(' > start - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        sr = Path(str(cfg[arg.section]['saved_filenames']))
        if arg.save_filenames is True and not sr.is_dir():
            ftemp = sr.open('w')
        else:
            ftemp = tempfile.TemporaryFile('w')
        z = str.maketrans(':', '_', '\\?()*$%&#@!`~\'"/<>{}[]|;,№^.')
        for i in xlsx_list:
            try:
                logger.info('  > {0}'.format(Path(i).name))
                r_book = openpyxl.load_workbook(str(i), read_only=True, use_iterators=True)
                r_sheet = r_book.active
                tmp_row = []
                cnt = 0
                cnt3= 0
                scount = 0
                header = []
                file_pool = {}
                sheet_pool = {}
                not_found_book = openpyxl.Workbook(optimized_write=True, guess_types=False, data_only=True)
                not_found_sheet = not_found_book.create_sheet(title='not_found')
                for rrr in r_sheet.rows:
                    if (cnt % 10000) == 0 and cnt != 0:
                        logger.info('  ...{0}'.format(str(cnt)))
                    if cnt == 0:
                        header = [string_none(i.value) for i in rrr]
                        cnt = cnt + 1
                        not_found_sheet.append(header)
                    else:
                        tmp_row = [string_none(i.value) for i in rrr]
                        xxx = mask.match(tmp_row[col_match])
                        if xxx is None:
                            not_found_sheet.append(tmp_row)
                            cnt = cnt + 1
                            cnt3 = cnt3 + 1
                        else:
                            result_file = str(Path(result_dir).joinpath(''.join([Path(Path(i).name).stem, '_', str(xxx.group(0)).translate(z), '.xlsx'])))
                            if result_file not in file_pool:
                                file_pool[result_file]  = openpyxl.Workbook(optimized_write=True, guess_types=False, data_only=True)
                                sheet_pool[result_file] = file_pool[result_file].create_sheet(title=''.join(['sheet', str(scount)]))
                                sheet_pool[result_file].append(header)
                            sheet_pool[result_file].append(tmp_row)
                            cnt = cnt + 1
                for j in file_pool:
                    file_pool[j].save(j)
                    ftemp.write(j+'\n')
                not_found_book.save(str(Path(result_dir).joinpath(''.join(['not_found_r', str(cnt3), '.xlsx']))))
                ftemp.write(str(Path(result_dir).joinpath(''.join(['not_found_r', str(cnt3), '.xlsx'])))+'\n')
            except Exception as ee:
                logger.error("{0} - ERROR: {1}\n{2}\n".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), i, ee))
                continue
        ftemp.close()
        logger.info(' > stop  - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
        return True
    else:
        logger.error('ERROR: No XLSX in {0} or unable to create {1}'.format(str(file_dir), str(result_dir)))
        return False

def XLSX_Compare(arg=None, cfg=None):
    logger.info('XLSX_Compare')
    to_file = Path(cfg[str(arg.section)]['source']).joinpath(cfg[str(arg.section)]['to_file'])
    from_file = Path(cfg[str(arg.section)]['source']).joinpath(cfg[str(arg.section)]['from_file'])
    result_dir = Path(cfg[str(arg.section)]['destination'])
    if not result_dir.exists():
        result_dir.mkdir(parents=True)
    if to_file.exists() and from_file.exists():
        try:
            sr = Path(str(cfg[arg.section]['saved_filenames']))
            if arg.save_filenames is True and not sr.is_dir():
                ftemp = sr.open('w')
            else:
                ftemp = tempfile.TemporaryFile('w')
            wb = openpyxl.Workbook(optimized_write = True, guess_types=False, data_only=True)
            scount = 0
            ws = wb.create_sheet(title=''.join(['sheet', str(scount)]))
            result_file = str(result_dir.joinpath(''.join([Path(Path(cfg[str(arg.section)]['to_file']).name).stem, '_END.xlsx'])))

            logger.info('loading file #1: {0}'.format(str(to_file)))
            logger.info('...')
            r_book_to = openpyxl.load_workbook(str(to_file), read_only=True, use_iterators=True)
            r_sheet_to = r_book_to.active
            to_content = [[j.value for j in y] for y in r_sheet_to.rows]

            logger.info('loading file #2: {0}'.format(str(from_file)))
            logger.info('...')
            r_book_from = openpyxl.load_workbook(str(from_file), read_only=True, use_iterators=True)
            r_sheet_from = r_book_from.active
            from_content = [[j.value for j in y] for y in r_sheet_from.rows]

            logger.info(' > start - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))

            cnt = 0
            empty_list = list('-'*len(from_content[0]))
            to_num = int(cfg[str(arg.section)]['to_col'])
            from_num = int(cfg[str(arg.section)]['from_col'])
            for ttt in to_content:
                tmp_ttt = [i for i in ttt]
                if (cnt % 1000) == 0 and cnt != 0:
                    logger.info('  ...{0}'.format(str(cnt)))
                is_found = 0
                cnt = cnt + 1
                for fff in from_content:
                    tmp_fff = [i for i in fff]
                    if (tmp_ttt[to_num] is not None and tmp_fff[from_num] is not None and (tmp_ttt[to_num].lower().strip() == tmp_fff[from_num].lower().strip())):
                        ws.append(tmp_ttt + ['|'] + tmp_fff)
                        is_found = 1
                        logger.info('  > {0}'.format(tmp_ttt[to_num]))
                if is_found == 0:
                    ws.append(tmp_ttt + ['|'] + empty_list)
            wb.save(result_file)
            ftemp.write(result_file+'\n')
            logger.info(' > stop  - {0}'.format(datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')))
            logger.info('{0}'.format(result_file))
            ftemp.close()
            return True
        except Exception as ee:
            logger.error("{0} - ERROR: {1}".format(datetime.datetime.now().strftime('%d %b %Y %H:%M:%S'), ee))
    else:
        logger.error('Please, check your config')
        return False

def ProcXLSX(arg=None, cfg=None):
    if arg and cfg:
        subtypes = {\
                'join': XLSX_Join,
                'split_rowcount': XLSX_Split_by_rowcount,
                'split_mask': XLSX_Split_by_mask,
                'compare': XLSX_Compare
                }

        subtypes[arg.type](arg, cfg)
        logger.info('END\n-----------\n\n')
        return True
    else:
        return False

if __name__ == '__main__':
    pass
